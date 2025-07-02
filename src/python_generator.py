"""
Python Code Generator

This module generates Python code equivalent to VBA macros.
"""

import logging
import re
from typing import Dict, List, Any, Optional
from vba_parser import VBAParser

logger = logging.getLogger(__name__)


class PythonGenerator:
    """Generate Python code from parsed VBA structure."""
    
    def __init__(self):
        self.parser = VBAParser()
        self.indent = "    "  # 4 spaces
        
        # VBA to Python type mapping
        self.type_mapping = {
            'INTEGER': 'int',
            'LONG': 'int',
            'STRING': 'str',
            'BOOLEAN': 'bool',
            'DOUBLE': 'float',
            'SINGLE': 'float',
            'VARIANT': 'Any',
            'DATE': 'datetime',
            'OBJECT': 'object',
            'CURRENCY': 'Decimal'
        }
        
        # VBA to Python function mapping
        self.function_mapping = {
            'MSGBOX': 'print',
            'LEN': 'len',
            'UCASE': 'str.upper',
            'LCASE': 'str.lower',
            'TRIM': 'str.strip',
            'LEFT': 'lambda s, n: s[:n]',
            'RIGHT': 'lambda s, n: s[-n:]',
            'MID': 'lambda s, start, length=None: s[start-1:start-1+length] if length else s[start-1:]'
        }
    
    def generate_python_code(self, vba_modules: List[Dict[str, str]]) -> str:
        """
        Generate Python code from VBA modules.
        
        Args:
            vba_modules: List of VBA modules with code
            
        Returns:
            Generated Python code string
        """
        python_code = []
        
        # Add imports
        python_code.extend(self._generate_imports())
        python_code.append("")
        
        # Process each module
        for module in vba_modules:
            module_name = module.get('module_name', module.get('name', 'Unknown'))
            python_code.append(f"# Converted from VBA module: {module_name}")
            python_code.append("")
            
            if module.get('type') == 'formulas':
                python_code.extend(self._convert_formulas_module(module.get('code', '')))
            else:
                python_code.extend(self._convert_vba_module(module.get('code', '')))
            
            python_code.append("")
        
        return '\n'.join(python_code)
    
    def _generate_imports(self) -> List[str]:
        """Generate necessary Python imports."""
        imports = [
            "import pandas as pd",
            "import openpyxl",
            "from openpyxl import Workbook, load_workbook",
            "from typing import Any, Optional, Union",
            "from datetime import datetime",
            "from decimal import Decimal",
            "import os",
            "import sys"
        ]
        return imports
    
    def _convert_vba_module(self, vba_code: str) -> List[str]:
        """Convert a VBA module to Python."""
        structure = self.parser.parse_code(vba_code)
        python_lines = []
        
        # Convert module-level variables
        if structure['variables']:
            python_lines.append("# Module-level variables")
            for var in structure['variables']:
                python_type = self.type_mapping.get(var['type'].upper(), 'Any')
                python_lines.append(f"{var['name'].lower()}: {python_type} = None")
            python_lines.append("")
        
        # Convert constants
        if structure['constants']:
            python_lines.append("# Constants")
            for const in structure['constants']:
                python_lines.append(f"{const['name'].upper()} = {const['value']}")
            python_lines.append("")
        
        # Convert functions
        for func in structure['functions']:
            python_lines.extend(self._convert_function(func, vba_code))
            python_lines.append("")
        
        return python_lines
    
    def _convert_function(self, func_info: Dict[str, Any], vba_code: str) -> List[str]:
        """Convert a VBA function to Python."""
        lines = vba_code.split('\n')
        func_lines = lines[func_info['start_line']-1:func_info['end_line']]
        
        python_lines = []
        
        # Generate function signature
        func_name = func_info['name'].lower()
        params = []
        
        for param in func_info['parameters']:
            param_name = param['name'].lower()
            param_type = self.type_mapping.get(param['type'].upper(), 'Any')
            
            if param['optional']:
                params.append(f"{param_name}: Optional[{param_type}] = None")
            else:
                params.append(f"{param_name}: {param_type}")
        
        # Determine return type
        return_type = "Any"
        if func_info['type'] == 'FUNCTION':
            return_type = "Any"  # Could be enhanced to detect return type
        else:
            return_type = "None"
        
        signature = f"def {func_name}({', '.join(params)}) -> {return_type}:"
        python_lines.append(signature)
        
        # Add docstring
        python_lines.append(f'{self.indent}"""')
        python_lines.append(f'{self.indent}Converted from VBA {func_info["type"]}: {func_info["name"]}')
        if func_info['parameters']:
            python_lines.append(f'{self.indent}')
            python_lines.append(f'{self.indent}Args:')
            for param in func_info['parameters']:
                python_lines.append(f'{self.indent}    {param["name"].lower()}: {param["type"]}')
        python_lines.append(f'{self.indent}"""')
        
        # Convert function body
        body_lines = self._convert_function_body(func_lines[1:-1])  # Exclude function declaration and end
        if not body_lines or all(not line.strip() for line in body_lines):
            python_lines.append(f"{self.indent}pass  # TODO: Implement function body")
        else:
            python_lines.extend(body_lines)
        
        return python_lines
    
    def _convert_function_body(self, vba_lines: List[str]) -> List[str]:
        """Convert VBA function body to Python."""
        python_lines = []
        indent_level = 1
        
        for line in vba_lines:
            stripped = line.strip()
            if not stripped or stripped.startswith("'"):
                if stripped.startswith("'"):
                    python_lines.append(f"{self.indent * indent_level}# {stripped[1:].strip()}")
                continue
            
            # Convert the line
            converted = self._convert_vba_line(stripped, indent_level)
            python_lines.extend(converted)
        
        return python_lines
    
    def _convert_vba_line(self, vba_line: str, indent_level: int) -> List[str]:
        """Convert a single VBA line to Python."""
        python_lines = []
        base_indent = self.indent * indent_level
        
        vba_upper = vba_line.upper()
        
        # Handle control structures
        if vba_upper.startswith('IF '):
            condition = self._convert_condition(vba_line[3:].replace(' THEN', ''))
            python_lines.append(f"{base_indent}if {condition}:")
            
        elif vba_upper.startswith('ELSEIF '):
            condition = self._convert_condition(vba_line[7:].replace(' THEN', ''))
            python_lines.append(f"{base_indent[:-4]}elif {condition}:")
            
        elif vba_upper.startswith('ELSE'):
            python_lines.append(f"{base_indent[:-4]}else:")
            
        elif vba_upper.startswith('END IF'):
            # Python doesn't need explicit end statements
            pass
            
        elif vba_upper.startswith('FOR '):
            loop_var, loop_range = self._parse_for_loop(vba_line)
            python_lines.append(f"{base_indent}for {loop_var} in {loop_range}:")
            
        elif vba_upper.startswith('NEXT'):
            # Python doesn't need explicit next statements
            pass
            
        elif vba_upper.startswith('WHILE '):
            condition = self._convert_condition(vba_line[6:])
            python_lines.append(f"{base_indent}while {condition}:")
            
        elif vba_upper.startswith('DIM '):
            var_declaration = self._convert_variable_declaration(vba_line)
            if var_declaration:
                python_lines.append(f"{base_indent}{var_declaration}")
        
        elif vba_upper.startswith('SET '):
            assignment = self._convert_set_statement(vba_line[4:])
            python_lines.append(f"{base_indent}{assignment}")
            
        else:
            # General statement conversion
            converted = self._convert_general_statement(vba_line)
            python_lines.append(f"{base_indent}{converted}")
        
        return python_lines
    
    def _convert_condition(self, condition: str) -> str:
        """Convert VBA condition to Python."""
        # Replace VBA operators with Python equivalents
        condition = condition.strip()
        condition = re.sub(r'\bAND\b', 'and', condition, flags=re.IGNORECASE)
        condition = re.sub(r'\bOR\b', 'or', condition, flags=re.IGNORECASE)
        condition = re.sub(r'\bNOT\b', 'not', condition, flags=re.IGNORECASE)
        condition = re.sub(r'<>', '!=', condition)
        
        return condition
    
    def _parse_for_loop(self, for_line: str) -> tuple:
        """Parse VBA FOR loop and convert to Python range."""
        # Example: FOR i = 1 TO 10 STEP 2
        match = re.search(r'FOR\s+(\w+)\s*=\s*(.+?)\s+TO\s+(.+?)(?:\s+STEP\s+(.+?))?(?:\s|$)', 
                         for_line, re.IGNORECASE)
        
        if match:
            var = match.group(1).lower()
            start = match.group(2)
            end = match.group(3)
            step = match.group(4) if match.group(4) else "1"
            
            return var, f"range({start}, {end} + 1, {step})"
        
        return "i", "range(10)  # TODO: Fix loop range"
    
    def _convert_variable_declaration(self, dim_line: str) -> Optional[str]:
        """Convert VBA DIM statement to Python variable annotation."""
        # Example: DIM myVar AS Integer
        match = re.search(r'DIM\s+(\w+)(?:\s+AS\s+(\w+))?', dim_line, re.IGNORECASE)
        
        if match:
            var_name = match.group(1).lower()
            var_type = match.group(2) if match.group(2) else 'Variant'
            python_type = self.type_mapping.get(var_type.upper(), 'Any')
            
            return f"{var_name}: {python_type} = None"
        
        return None
    
    def _convert_set_statement(self, set_line: str) -> str:
        """Convert VBA SET statement to Python assignment."""
        # Example: SET ws = ActiveSheet
        if '=' in set_line:
            var, value = set_line.split('=', 1)
            var = var.strip().lower()
            value = self._convert_expression(value.strip())
            return f"{var} = {value}"
        
        return set_line
    
    def _convert_general_statement(self, statement: str) -> str:
        """Convert general VBA statement to Python."""
        # Handle assignments
        if '=' in statement and not any(op in statement for op in ['<>', '<=', '>=']):
            var, value = statement.split('=', 1)
            var = var.strip().lower()
            value = self._convert_expression(value.strip())
            return f"{var} = {value}"
        
        # Handle method calls
        statement = self._convert_expression(statement)
        
        return statement
    
    def _convert_expression(self, expression: str) -> str:
        """Convert VBA expression to Python."""
        # Convert function calls
        for vba_func, python_func in self.function_mapping.items():
            pattern = r'\b' + vba_func + r'\s*\('
            if re.search(pattern, expression, re.IGNORECASE):
                expression = re.sub(pattern, f"{python_func}(", expression, flags=re.IGNORECASE)
        
        # Convert Excel object references
        expression = re.sub(r'\bRange\s*\(\s*"([^"]+)"\s*\)', r'worksheet["\1"]', expression, flags=re.IGNORECASE)
        expression = re.sub(r'\bCells\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)', r'worksheet.cell(\1, \2)', expression, flags=re.IGNORECASE)
        
        # Convert string concatenation
        expression = re.sub(r'\s*&\s*', ' + ', expression)
        
        return expression
    
    def _convert_formulas_module(self, formulas_code: str) -> List[str]:
        """Convert Excel formulas to Python pandas operations."""
        python_lines = [
            "class ExcelFormulasConverter:",
            "    \"\"\"Converted Excel formulas to Python operations\"\"\"",
            "    ",
            "    def __init__(self, workbook_path: str):",
            "        self.workbook_path = workbook_path",
            "        self.workbook = load_workbook(workbook_path)",
            "        self.data_frames = {}",
            "        self._load_sheets()",
            "    ",
            "    def _load_sheets(self):",
            "        \"\"\"Load all sheets as pandas DataFrames\"\"\"",
            "        for sheet_name in self.workbook.sheetnames:",
            "            ws = self.workbook[sheet_name]",
            "            data = []",
            "            for row in ws.iter_rows(values_only=True):",
            "                data.append(row)",
            "            self.data_frames[sheet_name] = pd.DataFrame(data)",
            "    ",
            "    def apply_formulas(self):",
            "        \"\"\"Apply converted formulas\"\"\"",
            "        # TODO: Implement formula conversions",
            "        pass"
        ]
        
        return python_lines