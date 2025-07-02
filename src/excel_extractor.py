"""
Excel VBA Extractor

This module provides functionality to extract VBA code and macros from Excel files.
"""

import logging
import os
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Any
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelExtractor:
    """Extract VBA code and data from Excel files."""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xlsm', '.xls']
    
    def extract_vba_code(self, file_path: Path) -> List[Dict[str, str]]:
        """
        Extract VBA code from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of dictionaries containing module name and code
        """
        try:
            if file_path.suffix.lower() == '.xlsm':
                return self._extract_from_xlsm(file_path)
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                return self._extract_from_xlsx(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")
        
        except Exception as e:
            logger.error(f"Error extracting VBA code from {file_path}: {e}")
            return []
    
    def _extract_from_xlsm(self, file_path: Path) -> List[Dict[str, str]]:
        """Extract VBA from macro-enabled Excel file (.xlsm)."""
        vba_modules = []
        
        try:
            # Extract VBA project from xlsm file
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                # Look for VBA project files
                vba_files = [f for f in zip_file.namelist() if 'vbaProject' in f]
                
                if not vba_files:
                    logger.warning(f"No VBA project found in {file_path}")
                    return []
                
                # Extract VBA modules (simplified approach)
                # In a real implementation, you'd use a VBA parser library
                for vba_file in vba_files:
                    try:
                        content = zip_file.read(vba_file)
                        # This is a simplified extraction - VBA is actually compiled
                        # For real VBA extraction, consider using oletools or similar
                        vba_modules.append({
                            'module_name': vba_file.split('/')[-1],
                            'code': self._parse_vba_content(content),
                            'type': 'module'
                        })
                    except Exception as e:
                        logger.warning(f"Could not extract {vba_file}: {e}")
        
        except Exception as e:
            logger.error(f"Error processing XLSM file: {e}")
        
        return vba_modules
    
    def _extract_from_xlsx(self, file_path: Path) -> List[Dict[str, str]]:
        """Extract any embedded code from regular Excel file."""
        # Regular xlsx files don't contain VBA, but might have formulas
        try:
            workbook = load_workbook(file_path, data_only=False)
            formulas = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'formula': cell.value
                            })
            
            if formulas:
                return [{
                    'module_name': 'Excel_Formulas',
                    'code': self._convert_formulas_to_vba_style(formulas),
                    'type': 'formulas'
                }]
        
        except Exception as e:
            logger.error(f"Error extracting from XLSX: {e}")
        
        return []
    
    def _parse_vba_content(self, content: bytes) -> str:
        """
        Parse VBA content from binary data.
        Note: This is a placeholder - real VBA parsing requires specialized libraries.
        """
        try:
            # VBA is stored in a compiled format, this is a simplified approach
            # For production use, consider libraries like oletools, python-oletools
            text_content = content.decode('utf-8', errors='ignore')
            
            # Extract readable text that might be VBA code
            lines = text_content.split('\n')
            vba_like_lines = []
            
            for line in lines:
                line = line.strip()
                if any(keyword in line.upper() for keyword in [
                    'SUB ', 'FUNCTION ', 'DIM ', 'IF ', 'THEN', 'ELSE',
                    'FOR ', 'NEXT', 'WHILE ', 'END SUB', 'END FUNCTION'
                ]):
                    vba_like_lines.append(line)
            
            return '\n'.join(vba_like_lines) if vba_like_lines else "' No readable VBA code found"
        
        except Exception as e:
            logger.warning(f"Could not parse VBA content: {e}")
            return "' VBA code extraction failed"
    
    def _convert_formulas_to_vba_style(self, formulas: List[Dict]) -> str:
        """Convert Excel formulas to VBA-style code for analysis."""
        vba_code = ["' Converted Excel Formulas to VBA-style code\n"]
        
        for formula in formulas:
            vba_code.append(f"' Sheet: {formula['sheet']}, Cell: {formula['cell']}")
            vba_code.append(f"Range(\"{formula['cell']}\").Formula = \"{formula['formula']}\"")
            vba_code.append("")
        
        return '\n'.join(vba_code)
    
    def get_vba_modules(self, file_path: Path) -> Dict[str, Any]:
        """Get metadata about VBA modules in the file."""
        vba_code = self.extract_vba_code(file_path)
        
        modules_info = {}
        for module in vba_code:
            modules_info[module['module_name']] = {
                'type': module['type'],
                'lines': len(module['code'].split('\n')),
                'functions': self._count_functions(module['code']),
                'variables': self._count_variables(module['code'])
            }
        
        return modules_info
    
    def _count_functions(self, code: str) -> int:
        """Count functions and subroutines in VBA code."""
        lines = code.upper().split('\n')
        count = 0
        for line in lines:
            line = line.strip()
            if line.startswith('SUB ') or line.startswith('FUNCTION '):
                count += 1
        return count
    
    def _count_variables(self, code: str) -> int:
        """Count variable declarations in VBA code."""
        lines = code.upper().split('\n')
        count = 0
        for line in lines:
            line = line.strip()
            if line.startswith('DIM '):
                count += 1
        return count
    
    def extract_workbook_structure(self, file_path: Path) -> Dict[str, Any]:
        """Extract the structure of the Excel workbook."""
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            structure = {
                'worksheets': [],
                'named_ranges': [],
                'properties': {}
            }
            
            # Get worksheet information
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                structure['worksheets'].append({
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'has_data': sheet.max_row > 1 or sheet.max_column > 1
                })
            
            # Get named ranges if any
            if hasattr(workbook, 'defined_names'):
                for name in workbook.defined_names:
                    structure['named_ranges'].append({
                        'name': name.name,
                        'reference': str(name.attr_text) if hasattr(name, 'attr_text') else 'Unknown'
                    })
            
            return structure
        
        except Exception as e:
            logger.error(f"Error extracting workbook structure: {e}")
            return {}